from datetime import date
from pathlib import Path
from typing import cast

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


COLUMNS = [
    "query_date",
    "search_keyword",
    "timestamp",
    "account",
    "link",
    "tag",
    "content",
]
HEADERS = {
    "query_date": "查詢日期",
    "search_keyword": "搜尋關鍵字",
    "timestamp": "發文時間",
    "account": "帳號",
    "link": "貼文連結",
    "tag": "標籤",
    "content": "內容",
}
COL_WIDTHS = {
    "query_date": 14,
    "search_keyword": 20,
    "timestamp": 24,
    "account": 20,
    "link": 50,
    "tag": 30,
    "content": 60,
}


def _init_workbook() -> tuple[openpyxl.Workbook, Worksheet]:
    wb = openpyxl.Workbook()
    ws = cast(Worksheet, wb.active)
    ws.title = "Threads 貼文"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4A4A4A", end_color="4A4A4A", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, col_key in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=HEADERS[col_key])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[col_key]

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    return wb, ws


def _existing_links(ws, link_col_idx: int) -> set[str]:
    """讀取工作表中所有已存在的貼文連結（用於去重）。"""
    links: set[str] = set()
    for row in ws.iter_rows(
        min_row=2, min_col=link_col_idx, max_col=link_col_idx, values_only=True
    ):
        if row[0]:
            links.add(str(row[0]))
    return links


def export(posts: list[dict], output_path: str) -> Path:
    """將貼文列表匯出為 Excel；若檔案已存在則 append 並以連結去重。"""
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    query_date = date.today().isoformat()
    link_col_idx = COLUMNS.index("link") + 1  # 1-based

    # 載入既有檔案或建立新檔
    if output.exists():
        wb = openpyxl.load_workbook(str(output))
        ws = cast(Worksheet, wb.active)
        existing = _existing_links(ws, link_col_idx)
        start_row = ws.max_row + 1
        print(f"檔案已存在，目前有 {ws.max_row - 1} 筆資料，將以 append 模式追加")
    else:
        wb, ws = _init_workbook()
        existing = set()
        start_row = 2

    # 過濾已存在的貼文
    new_posts = [p for p in posts if p.get("link", "") not in existing]
    if len(new_posts) < len(posts):
        print(
            f"去重：{len(posts)} 篇 → {len(new_posts)} 篇新貼文（跳過 {len(posts) - len(new_posts)} 篇重複）"
        )

    # 寫入資料行
    wrap = Alignment(wrap_text=True, vertical="top")
    for row_idx, post in enumerate(new_posts, start=start_row):
        row_data = [
            query_date,
            post.get("search_keyword", ""),
            post.get("timestamp", ""),
            post.get("account", ""),
            post.get("link", ""),
            post.get("tag", ""),
            post.get("content", ""),
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = wrap

    wb.save(str(output))
    return output
